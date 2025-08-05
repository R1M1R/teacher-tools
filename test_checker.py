import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

def check_tests(input_file, output_file):
    """
    Проверяет тесты и создает файл с результатами
    :param input_file: Файл с ответами (xlsx)
    :param output_file: Файл для сохранения результатов
    """
    try:
        # Чтение данных
        df = pd.read_excel(input_file)
        
        # Проверка наличия нужных колонок
        required_columns = ['ФИО', 'Ответ ученика', 'Правильный ответ']
        if not all(col in df.columns for col in required_columns):
            raise ValueError("Неверный формат файла! Нужны колонки: ФИО, Ответ ученика, Правильный ответ")

        # Проверка ответов
        df['Результат'] = df['Ответ ученика'] == df['Правильный ответ']
        df['Балл'] = df['Результат'].astype(int)
        
        # Сохранение в Excel с подсветкой
        df.to_excel(output_file, index=False)
        
        # Форматирование Excel
        wb = load_workbook(output_file)
        ws = wb.active
        
        red_fill = PatternFill(start_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", fill_type="solid")
        
        for row in range(2, len(df)+2):
            if ws.cell(row=row, column=3).value != ws.cell(row=row, column=4).value:
                ws.cell(row=row, column=3).fill = red_fill
            else:
                ws.cell(row=row, column=3).fill = green_fill
        
        wb.save(output_file)
        
        # Статистика
        total = len(df)
        correct = df['Балл'].sum()
        return f"Проверено {total} работ. Успешно: {correct} ({correct/total:.0%})"
        
    except Exception as e:
        return f"Ошибка: {str(e)}"

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Использование: python test_checker.py входной_файл.xlsx выходной_файл.xlsx")
    else:
        result = check_tests(sys.argv[1], sys.argv[2])
        print(result)
